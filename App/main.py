import tkinter as tk
from tkinter import filedialog, Label, Button, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from PIL import Image, ImageTk
import pandas as pd
import re
import os
# import sys
# import random
# import xlwings as xw
import numpy as np

pd.set_option('future.no_silent_downcasting', True)

color_dict = {
    'Y': 'FFFF00',  # Yellow
    'G': '00FF00',  # Green
    'L': '0000FF',  # Blue
    'R': 'FF0000',  # Red
    'W': 'FFFFFF',  # White
    'LG': '808080', # Gray (Light Gray)
    'O': 'FFA500',  # Orange
    'BR': '80471c',#'A52A2A', # Brown
    'V': '800080',#'8A2BE2',   # Violet
    'GY': '707070',#'C0C0C0',  # Gray
    'B': '000000' ,  # Black (Noir)
    'P' : 'FFC0CB', # Pink
    'C' : '00FFFF', # Cyan
    'D' : 'FFFFF0', # Ivory
    'S' : 'C0C0C0', # Silver
}

def apply_color(symbol, cell):
    if '/' in str(symbol):
        primary, secondary = symbol.split('/')
        fill = PatternFill(start_color=color_dict[primary], end_color=color_dict[primary], fill_type="solid")
        cell.fill = fill
        
        side = Side(border_style="thick", color=color_dict[secondary])
        border = Border(diagonal=side, diagonalUp=True)
        cell.border = border
    elif symbol in color_dict:
        fill = PatternFill(start_color=color_dict[symbol], end_color=color_dict[symbol], fill_type="solid")
        cell.fill = fill
    else :
        cell.value = symbol
    
# def header_final_schema(ws):
#     # wb = load_workbook(file_path)
#     # ws = wb.active 
#     print('header_final_schema')
#     print(ws)
#     print(f"Sheet Title : {ws.title}")
#     column_widths = [15, 20, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
#     for i, width in enumerate(column_widths, 1):
#         ws.column_dimensions[get_column_letter(i)].width = width
    
#     thin_border = Border(
#         left=Side(style='medium'), 
#         right=Side(style='medium'), 
#         top=Side(style='medium'), 
#         bottom=Side(style='medium')
#     )
    
#     # Header row 1 - Yazaki logo area
#     img = Image('../yazaki_logo.png')
#     ws.merge_cells('A1:J1')
#     cell_logo = 'A2'
#     ws.add_image(img, cell_logo)
#     # ws['A1'] = 'YAZAKI'
#     # ws['A1'].font = Font(name='Arial', size=14, bold=True)
#     # ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    
#     blue_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
#     blue_font = Font(name='Arial', size=10, color=Color(rgb='FFFFFF'))
    
#     ws.merge_cells('K1:O1')
#     first_cell = ws.cell(row=1, column=11)  
#     first_cell.value = 'PU24/PU25 LHD'
#     first_cell.font = blue_font
#     first_cell.fill = blue_fill
#     first_cell.alignment = Alignment(horizontal='center', vertical='center')
    
#     header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
#     header_font = Font(name='Arial', size=10)
    
#     cell_configs = [
#         ('C2', 'Nº du Produit / Niveau', header_font, header_fill, thin_border),
#         ('C3', 'KAR G60', Font(name='Arial', size=10), None, thin_border),
#         ('E2', 'Local du travail', header_font, header_fill, thin_border),
#         ('E3', 'WP414-SA513', Font(name='Arial', size=10), None, thin_border),
#         ('F2', 'N° de ligne', header_font, header_fill, thin_border),
#         ('F3', '1', Font(name='Arial', size=10), None, thin_border),
#         ('G2', 'N° de ligne', header_font, header_fill, thin_border),
#         ('G3', '1', Font(name='Arial', size=10), None, thin_border),
#         ('H2', 'N° de ligne', header_font, header_fill, thin_border),
#         ('H3', '1', Font(name='Arial', size=10), None, thin_border),
#         ('I2:J2', 'Processus', header_font, header_fill, thin_border),
#         ('I3:J3', 'C2', Font(name='Arial', size=10), None, thin_border),
#         ('M2', 'Nº de Registre', header_font, header_fill, thin_border),
#         ('M3', 'EA-EN-MMO-xx-T-6047', Font(name='Arial', size=10), None, thin_border)
#     ]
    
#     for cell_range, value, font, fill, border in cell_configs:
#         if ':' in cell_range:
#             ws.merge_cells(cell_range)
#             # Get the first cell of the merged range
#             first_cell = ws[cell_range.split(':')[0]]
#             first_cell.value = value
#             first_cell.font = font
            
#             if fill:
#                 first_cell.fill = fill
            
#             first_cell.border = border
#             first_cell.alignment = Alignment(horizontal='center', vertical='center')
#         else:
#             cell = ws[cell_range]
#             cell.value = value
#             cell.font = font
            
#             if fill:
#                 cell.fill = fill
            
#             cell.border = border
#             cell.alignment = Alignment(horizontal='center', vertical='center')
    
#     # Add borders to entire merged ranges
#     border_ranges = ['C2:D2', 'I2:J2', 'M2:O2', 'C3:D3', 'I3:J3', 'M3:O3']
#     for cell_range in border_ranges:
#         for row in ws[cell_range]:
#             for cell in row:
#                 cell.border = thin_border
    
#     print("Yazaki header template created successfully!")
#     # wb.save(file_path)
#     return ws

def header_final_schema(file_path):
    # wb = load_workbook(file_path)
    # ws = wb.active 
    
    wb = load_workbook(file_path)
    ws = wb.active
    for ws in wb.worksheets:
        column_widths = [15, 20, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        thin_border = Border(
            left=Side(style='medium'), 
            right=Side(style='medium'), 
            top=Side(style='medium'), 
            bottom=Side(style='medium')
        )
        
        # Header row 1 - Yazaki logo area
        img = Image('../yazaki_logo.png')
        cell_logo = ws.merge_cells('A1:J1')
        ws.add_image(img, cell_logo)
        # ws['A1'] = 'YAZAKI'
        # ws['A1'].font = Font(name='Arial', size=14, bold=True)
        # ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
        blue_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        blue_font = Font(name='Arial', size=10, color=Color(rgb='FFFFFF'))
        
        ws.merge_cells('K1:O1')
        first_cell = ws.cell(row=1, column=11)  
        first_cell.value = 'PU24/PU25 LHD'
        first_cell.font = blue_font
        first_cell.fill = blue_fill
        first_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        header_font = Font(name='Arial', size=10)
        
        cell_configs = [
            ('C2', 'Nº du Produit / Niveau', header_font, header_fill, thin_border),
            ('C3', 'KAR G60', Font(name='Arial', size=10), None, thin_border),
            ('E2', 'Local du travail', header_font, header_fill, thin_border),
            ('E3', 'WP414-SA513', Font(name='Arial', size=10), None, thin_border),
            ('F2', 'N° de ligne', header_font, header_fill, thin_border),
            ('F3', '1', Font(name='Arial', size=10), None, thin_border),
            ('G2', 'N° de ligne', header_font, header_fill, thin_border),
            ('G3', '1', Font(name='Arial', size=10), None, thin_border),
            ('H2', 'N° de ligne', header_font, header_fill, thin_border),
            ('H3', '1', Font(name='Arial', size=10), None, thin_border),
            ('I2:J2', 'Processus', header_font, header_fill, thin_border),
            ('I3:J3', 'C2', Font(name='Arial', size=10), None, thin_border),
            ('M2', 'Nº de Registre', header_font, header_fill, thin_border),
            ('M3', 'EA-EN-MMO-xx-T-6047', Font(name='Arial', size=10), None, thin_border)
        ]
        
        for cell_range, value, font, fill, border in cell_configs:
            if ':' in cell_range:
                ws.merge_cells(cell_range)
                # Get the first cell of the merged range
                first_cell = ws[cell_range.split(':')[0]]
                first_cell.value = value
                first_cell.font = font
                
                if fill:
                    first_cell.fill = fill
                
                first_cell.border = border
                first_cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell = ws[cell_range]
                cell.value = value
                cell.font = font
                
                if fill:
                    cell.fill = fill
                
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add borders to entire merged ranges
        border_ranges = ['C2:D2', 'I2:J2', 'M2:O2', 'C3:D3', 'I3:J3', 'M3:O3']
        for cell_range in border_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = thin_border
        
        print("Yazaki header template created successfully!")
    wb.save(file_path)

def schema_final_without_color(path_output_excel,df_input,all_sheets) :
    with pd.ExcelWriter(path_output_excel, engine='xlsxwriter') as writer:      
        for sheet in all_sheets:
            if pd.isna(sheet):
                continue
            
            sps_data = df_input[df_input['SPS'] == sheet]
            
            if sps_data.empty:
                continue
            
            pm = sps_data['Production Module MA15'].fillna('').astype(str)
            step = sps_data['Step'].fillna('').infer_objects(copy=False).astype(str)
            
            arrays = [np.array(pm), np.array(step)]
            index = pd.MultiIndex.from_arrays(arrays, names=('Production Module MA15', 'Sequence'))
            
            # Création du DataFrame avec la colonne renommée
            df = pd.DataFrame({'Materiel': sps_data['Component Name'].values,
                            'SAP NO MA15': sps_data['SAP NO MA15'].values,
                            'Note': sps_data['Note'].values,
                            'CS' : sps_data['CS'].values,
                            'color' : sps_data['Colour'].values,
                            '' : '',
                            'CON A' : sps_data['From Connector'].values,
                            'CAV A' : sps_data['From Cavity'].values,
                            'INSERTION A' : '',
                            'CON B' : sps_data['To Connector'].values,
                            'CAV B' : sps_data['To Cavity'].values,
                            'INSERTION B' : '',
                            },
                            index=index)
            df.to_excel(writer, sheet_name=sheet, startrow=4,startcol=0,index=True)
    print('Schema without color created successfully!')

def add_unique_ids_to_schema(wb,sheet_name):
    
    ws = wb[sheet_name]
        
    header_row = 5
        
    ws.insert_cols(1)
    id_cell = ws.cell(row=header_row, column=1)
    id_cell.value = "Séquence PM"
    id_cell.font = Font(bold=True)
    id_cell.alignment = Alignment(horizontal='center', vertical='center')
        
    id_counter = 1
    for row in range(header_row + 1, ws.max_row + 1):
        has_content = any(ws.cell(row=row, column=col).value for col in range(2, ws.max_column + 1))
        if has_content:
            id_cell = ws.cell(row=row, column=1)
            id_cell.value = id_counter
            id_cell.alignment = Alignment(horizontal='center', vertical='center')
            id_counter += 1
    

    return ws

def schema_final_with_color(wb_path):
    wb = load_workbook(wb_path)

    for sheet_name in wb.sheetnames:
        
        sheet = add_unique_ids_to_schema(wb,sheet_name)
        
        color_col = None
        header_row = 5
        
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=header_row, column=col).value == "color":
                color_col = col
                break
        
        if color_col is None:
            print(f"No color column found in sheet: {sheet_name}")
            continue
        
        # Add new X column header after last column
        # x_col = sheet.max_column + 1
        # sheet.cell(row=header_row, column=x_col).value = "X"
        
        for row in range(header_row + 1, sheet.max_row + 1):
            color_symbol = sheet.cell(row=row, column=color_col).value
            if color_symbol:
                cell = sheet.cell(row=row, column=color_col + 1)
                try:
                    apply_color(color_symbol, cell)
                except KeyError:
                    print(f"Unknown color code: {color_symbol} in sheet {sheet_name}, row {row}")
                    cell.value = "Unknown color code" #PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    print("Color formatting completed successfully.")
    wb.save(wb_path)

def add_last_table(file_path):
    print('add last table')
    wb = load_workbook(file_path)
    ws = wb.active 
    
    for ws in wb.worksheets:
        row_start = ws.max_row + 2  
        table_data = [
            ["", "PM Basique", "", "Niveau", "N° de Phase", "Date", "Préparé par", "Timbre"],
            ["Note:", "Les cases colorées sont des PM optionnelles", "", "", "", "", "", ""],
            ["", "🌀 : ", "A Insérer", "", "", "", "", ""],
            ["", "Ø : ", "A Ne pas insérer", "", "", "", "", ""],
            ["", "⊖ : ", "Déjà inséré", "", "", "", "", ""]
        ]

        # Définir les bordures
        border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )


        for row_index, row_data in enumerate(table_data, start=row_start):
            for col_index, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index, value=cell_value)
                ws.cell(row=row_index, column=col_index).border = border
                ws.cell(row=row_index, column=col_index).font = Font(size=12, bold=True)



            
    # # Appliquer les bordures        
    # for row in ws.iter_rows(min_row=row_start, max_row=row_start + len(table_data) - 1, min_col=1, max_col=len(table_data[0])):
    #     for cell in row:
    #         cell.border = border
    print('done...')        
    # print("Tableau ajouté avec succès !")
    wb.save(file_path)

def separator(input_file, path_output_excel):
    df_input = pd.read_excel(input_file)
    all_sheets = list(df_input['SPS'].unique())
    PM_sps = list(df_input['Production Module MA15'].unique())
    
    schema_final_without_color(path_output_excel,df_input,all_sheets)
    header_final_schema(path_output_excel)
    schema_final_with_color(path_output_excel)
    
    add_last_table(path_output_excel)
    

            
def select_ref_file():
    global Schema_file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls;*.XLSX")])
    
    if not file_path.lower().endswith((".xlsx", ".xls")):
        messagebox.showerror("Erreur", "Veuillez sélectionner un fichier Excel valide (*.xlsx, *.xls;*.XLSX)")
        return

    Schema_file_path = file_path
    Schema_label.config(text=f"Schema File: {Schema_file_path}") 

def select_output_dir():
    global output_dir
    output_dir = filedialog.askdirectory()
    output_label.config(text=f"Dossier de sortie: {output_dir}")

def main_separator():
    if not Schema_file_path or not output_dir:
        messagebox.showerror("Erreur", "Veuillez sélectionner le fichier Schéma et le dossier de sortie")
        return 
    
    status_text = f"Schema file: {Schema_file_path}\n"   
    if output_dir:
        status_text += f"\nOutput folder: {output_dir}" 
        
    try:
        output_file = f"{output_dir}/Final_Schema.xlsx"
        print("Séparation des fichiers...")

        messagebox.showinfo("showinfo", "Séparation des fichiers en cours...")
        separator(Schema_file_path,output_file)
       
        messagebox.showinfo("Succès", "Séparation des fichiers terminée avec succès !")
        return output_file
    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur lors de la Séparation des fichiers: {e}")
        return


root = tk.Tk()
root.title("Excel Comparator - YMM-2")
root.geometry("200x200")

# window_width = 300
# window_height = 300

# # Obtenir les dimensions de l'écran
# screen_width = root.winfo_screenwidth()
# screen_height = root.winfo_screenheight()

# # Calculer la position x et y
# x_position = (screen_width - window_width) // 2
# y_position = (screen_height - window_height) // 2

# # Appliquer la position centrée
# root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")

# # logo_path = "yazaki_logo.png"  
# try:
#     img = Image.open(logo_path)
#     img = img.resize((150, 50), Image.LANCZOS)
#     logo = ImageTk.PhotoImage(img)
#     root.logo = logo
#     logo_label = Label(root, image=logo)
#     logo_label.pack(pady=10)
# except:
#     logo_label = Label(root, text="[Logo Yazaki]", font=("Arial", 14, "bold"))
#     logo_label.pack(pady=10)


Schema_file_path = ""
output_dir = ""

Schema_button = Button(root, text="Schema", command=select_ref_file)
Schema_button.pack(pady=5)
Schema_label = Label(root, text="Référence: Non sélectionné", wraplength=400)
Schema_label.pack()


output_button = Button(root, text="Sélectionner le dossier de sortie", command=select_output_dir)
output_button.pack(pady=5)
output_label = Label(root, text="Dossier de sortie: Non sélectionné", wraplength=400)
output_label.pack()

compare_button = Button(root, text="Comparer", command=main_separator)
compare_button.pack(pady=20)

status_label = Label(root, text="", fg="black")
status_label.pack()

tk.mainloop()