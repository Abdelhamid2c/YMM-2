import tkinter as tk
from tkinter import filedialog, Label, Button, messagebox
from PIL import Image, ImageTk
import re
import os
import sys
# import random
# import xlwings as xw
import numpy as np

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import time

pd.set_option('future.no_silent_downcasting', True)

color_dict = {
    'Y': 'FFFF00',  # Yellow
    'G': '00FF00',  # Green
    'L': '0000FF',  # Blue
    'R': 'FF0000',  # Red
    'W': 'FFFFFF',  # White
    'LG': '808080', # Gray (Light Gray)
    'O': 'FFA500',  # Orange
    'BR': '80471c',#'A52A2A', # Brown
    'V': '800080',#'8A2BE2',   # Violet
    'GY': '707070',#'C0C0C0',  # Gray
    'B': '000000' ,  # Black (Noir)
    'P' : 'FFC0CB', # Pink
    'C' : '00FFFF', # Cyan
    'D' : 'FFFFF0', # Ivory
    'SI' : '505050', # Silver
}


def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)


def apply_color(symbol, cell):
    if '/' in str(symbol):
        primary, secondary = symbol.split('/')
        fill = PatternFill(start_color=color_dict[primary], end_color=color_dict[primary], fill_type="solid")
        cell.fill = fill
        
        side = Side(border_style="thick", color=color_dict[secondary])
        border = Border(diagonal=side, diagonalUp=True)
        cell.border = border
    elif symbol in color_dict:
        fill = PatternFill(start_color=color_dict[symbol], end_color=color_dict[symbol], fill_type="solid")
        cell.fill = fill
    else :
        cell.value = symbol
    

def header_final_schema(file_path):
    try :
        # wb = load_workbook(file_path)
        # ws = wb.active 
        
        wb = load_workbook(file_path)
        ws = wb.active
        for ws in wb.worksheets:
            column_widths = [15, 20, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            thin_border = Border(
                left=Side(style='medium'), 
                right=Side(style='medium'), 
                top=Side(style='medium'), 
                bottom=Side(style='medium')
            )
            
            # Header row 1 - Yazaki logo area
            # img = Image('../yazaki_logo.png')
            logo_path = resource_path("yazaki_logo.png")
            # cell_logo = ws.merge_cells('A1:J1')
            # ws.add_image(img, cell_logo)
            if os.path.exists(logo_path):
                try:
                    img = Image(logo_path)
                    ws.merge_cells('A1:J1')
                    ws.add_image(img, 'A2')
                except Exception as e:
                    print(f"Could not add logo: {e}")
                    ws['A1'] = 'YAZAKI'
                    ws['A1'].font = Font(name='Arial', size=14, bold=True)
                    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
            else:
                print(f"Logo file not found at {logo_path}")
                ws['A1'] = 'YAZAKI'
                ws['A1'].font = Font(name='Arial', size=14, bold=True)
                ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
            # ws['A1'] = 'YAZAKI'
            # ws['A1'].font = Font(name='Arial', size=14, bold=True)
            # ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
            
            blue_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
            blue_font = Font(name='Arial', size=10, color=Color(rgb='FFFFFF'))
            
            ws.merge_cells('K1:O1')
            first_cell = ws.cell(row=1, column=11)  
            first_cell.value = 'PU24/PU25 LHD'
            first_cell.font = blue_font
            first_cell.fill = blue_fill
            first_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            header_font = Font(name='Arial', size=10)
            
            cell_configs = [
                ('C2', 'Nº du Produit / Niveau', header_font, header_fill, thin_border),
                ('C3', 'KAR G60', Font(name='Arial', size=10), None, thin_border),
                ('E2', 'Local du travail', header_font, header_fill, thin_border),
                ('E3', 'WP414-SA513', Font(name='Arial', size=10), None, thin_border),
                ('F2', 'N° de ligne', header_font, header_fill, thin_border),
                ('F3', '1', Font(name='Arial', size=10), None, thin_border),
                ('G2', 'N° de ligne', header_font, header_fill, thin_border),
                ('G3', '1', Font(name='Arial', size=10), None, thin_border),
                ('H2', 'N° de ligne', header_font, header_fill, thin_border),
                ('H3', '1', Font(name='Arial', size=10), None, thin_border),
                ('I2:J2', 'Processus', header_font, header_fill, thin_border),
                ('I3:J3', 'C2', Font(name='Arial', size=10), None, thin_border),
                ('M2', 'Nº de Registre', header_font, header_fill, thin_border),
                ('M3', 'EA-EN-MMO-xx-T-6047', Font(name='Arial', size=10), None, thin_border)
            ]
            
            for cell_range, value, font, fill, border in cell_configs:
                if ':' in cell_range:
                    ws.merge_cells(cell_range)
                    # Get the first cell of the merged range
                    first_cell = ws[cell_range.split(':')[0]]
                    first_cell.value = value
                    first_cell.font = font
                    
                    if fill:
                        first_cell.fill = fill
                    
                    first_cell.border = border
                    first_cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell = ws[cell_range]
                    cell.value = value
                    cell.font = font
                    
                    if fill:
                        cell.fill = fill
                    
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add borders to entire merged ranges
            border_ranges = ['C2:D2', 'I2:J2', 'M2:O2', 'C3:D3', 'I3:J3', 'M3:O3']
            for cell_range in border_ranges:
                for row in ws[cell_range]:
                    for cell in row:
                        cell.border = thin_border
            
            # print("Yazaki header template created successfully!")
        wb.save(file_path)
        print("Header created successfully!")
        return True
    except Exception as e:
        print(f"Error in header_final_schema: {e}")
        return False

def schema_final_without_color(path_output_excel,df_input,all_sheets) :
    try:
        with pd.ExcelWriter(path_output_excel, engine='xlsxwriter') as writer:      
            for sheet in all_sheets:
                if pd.isna(sheet):
                    continue
                
                sps_data = df_input[df_input['SPS'] == sheet]
                
                if sps_data.empty:
                    continue
                
                pm = sps_data['Production Module MA15'].fillna('').astype(str)
                step = sps_data['Step'].fillna('').infer_objects(copy=False).astype(str)
                
                arrays = [np.array(pm), np.array(step)]
                index = pd.MultiIndex.from_arrays(arrays, names=('Production Module MA15', 'Sequence'))
                
                # Création du DataFrame avec la colonne renommée
                df = pd.DataFrame({'Materiel': sps_data['Component Name'].values,
                                'SAP NO MA15': sps_data['SAP NO MA15'].values,
                                'Note': sps_data['Note'].values,
                                'CS' : sps_data['CS'].values,
                                'color' : sps_data['Colour'].values,
                                '' : '',
                                'CON A' : sps_data['From Connector'].values,
                                'CAV A' : sps_data['From Cavity'].values,
                                'INSERTION A' : '',
                                'CON B' : sps_data['To Connector'].values,
                                'CAV B' : sps_data['To Cavity'].values,
                                'INSERTION B' : '',
                                },
                                index=index)
                df.to_excel(writer, sheet_name=sheet, startrow=4,startcol=0,index=True)
        print('Schema without color created successfully!')
        return True
    except Exception as e:
        print(f"Error in schema_final_without_color: {e}")
        return False


def add_unique_ids_to_schema(wb,sheet_name):
    
    ws = wb[sheet_name]
        
    header_row = 5
        
    ws.insert_cols(1)
    id_cell = ws.cell(row=header_row, column=1)
    id_cell.value = "Séquence PM"
    id_cell.font = Font(bold=True)
    id_cell.alignment = Alignment(horizontal='center', vertical='center')
        
    id_counter = 1
    for row in range(header_row + 1, ws.max_row + 1):
        has_content = any(ws.cell(row=row, column=col).value for col in range(2, ws.max_column + 1))
        if has_content:
            id_cell = ws.cell(row=row, column=1)
            id_cell.value = id_counter
            id_cell.alignment = Alignment(horizontal='center', vertical='center')
            id_counter += 1
    

    return ws

def schema_final_with_color(wb_path):
    try :
        wb = load_workbook(wb_path)

        for sheet_name in wb.sheetnames:
            
            sheet = add_unique_ids_to_schema(wb,sheet_name)
            
            color_col = None
            header_row = 5
            
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=header_row, column=col).value == "color":
                    color_col = col
                    break
            
            if color_col is None:
                print(f"No color column found in sheet: {sheet_name}")
                continue
            
            # Add new X column header after last column
            # x_col = sheet.max_column + 1
            # sheet.cell(row=header_row, column=x_col).value = "X"
            
            for row in range(header_row + 1, sheet.max_row + 1):
                color_symbol = sheet.cell(row=row, column=color_col).value
                if color_symbol:
                    cell = sheet.cell(row=row, column=color_col + 1)
                    try:
                        apply_color(color_symbol, cell)
                    except KeyError:
                        print(f"Unknown color code: {color_symbol} in sheet {sheet_name}, row {row}")
                        cell.value = "Unknown color code" #PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        print("Color formatting completed successfully.")
        wb.save(wb_path)
        return True
    except Exception as e:
        print(f"Error in schema_final_with_color: {e}")
        return False

def add_last_table(file_path):
    try :
        print('add last table')
        wb = load_workbook(file_path)
        ws = wb.active 
        
        for ws in wb.worksheets:
            row_start = ws.max_row + 2  
            table_data = [
                ["", "PM Basique", "", "Niveau", "N° de Phase", "Date", "Préparé par", "Timbre"],
                ["Note:", "Les cases colorées sont des PM optionnelles", "", "", "", "", "", ""],
                ["", "🌀 : ", "A Insérer", "", "", "", "", ""],
                ["", "Ø : ", "A Ne pas insérer", "", "", "", "", ""],
                ["", "⊖ : ", "Déjà inséré", "", "", "", "", ""]
            ]

            # Définir les bordures
            border = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )


            for row_index, row_data in enumerate(table_data, start=row_start):
                for col_index, cell_value in enumerate(row_data, start=1):
                    ws.cell(row=row_index, column=col_index, value=cell_value)
                    ws.cell(row=row_index, column=col_index).border = border
                    ws.cell(row=row_index, column=col_index).font = Font(size=12, bold=True)



                
        # # Appliquer les bordures        
        # for row in ws.iter_rows(min_row=row_start, max_row=row_start + len(table_data) - 1, min_col=1, max_col=len(table_data[0])):
        #     for cell in row:
        #         cell.border = border
        print('done...')        
        # print("Tableau ajouté avec succès !")
        wb.save(file_path)
        return True
    except Exception as e:
        print(f"Error in add_last_table: {e}")
        return False

def select_ref_file():
    global Schema_file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls;*.XLSX")])
    
    if not file_path.lower().endswith((".xlsx", ".xls")):
        messagebox.showerror("Erreur", "Please select a valid Excel file (*.xlsx, *.xls;*.XLSX)")
        return

    Schema_file_path = file_path
    Schema_label.config(text=f"Schema File: {Schema_file_path}") 

def select_output_dir():
    global output_dir
    output_dir = filedialog.askdirectory()
    output_label.config(text=f"Output folder: {output_dir}")

def main_separator():
    if not Schema_file_path or not output_dir:
        messagebox.showerror("Erreur", "Please select both Schema file and output folder")
        return 
    
    status_text = f"Schema file: {Schema_file_path}\n"   
    if output_dir:
        status_text += f"\nOutput folder: {output_dir}" 
        
    try:
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        output_file = os.path.join(output_dir, f"Final_Schema_{timestamp}.xlsx")
        # output_file = f"{output_dir}/Final_Schema.xlsx"
        compare_button.config(state=tk.DISABLED)
        root.update_idletasks()
        success = separator(Schema_file_path,output_file)
        compare_button.config(state="normal")
        if success:
            status_label.config(text=f"Success! File saved as Final_Schema_{timestamp}.xlsx", fg="green")
            messagebox.showinfo("Success", "File processing completed successfully!")
        else:
            status_label.config(text="Processing failed.", fg="red")
        
        # messagebox.showinfo("Succès", "File separation completed successfully!")
        # status_text += f"\nFichiers séparés avec succès !"
        # status_label.config(text=status_text, fg="green")
    except Exception as e:
        messagebox.showerror("Erreur", f"Error during file separation: {e}")
        return


def separator(file_path, output_file):
    try:
        df_input = pd.read_excel(file_path)
        all_sheets = list(df_input['SPS'].unique())
        PM_sps = list(df_input['Production Module MA15'].unique())
        
        if not all_sheets:
            messagebox.showerror("Error", "No valid SPS values found in the input file")
            return False
        
        
        if not schema_final_without_color(output_file,df_input,all_sheets):
            return False
            
        if not header_final_schema(output_file):
            return False
            
        if not schema_final_with_color(output_file):
            return False
        
        if not add_last_table(output_file):
            return False
        
        if not header_final_schema(output_file):
            return False
        
        print(f"Workbook saved successfully to {output_file}")
        return True
        
    except Exception as e:
        print(f"Error: {e}")
        messagebox.showerror("Error", f"Error processing the file: {e}")
        return False

root = tk.Tk()
root.title("Excel Schema Processor - YMM-2")
root.geometry("400x300")
window_width = 400
window_height = 300
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x_position = (screen_width - window_width) // 2
y_position = (screen_height - window_height) // 2
root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")

title_label = Label(root, text="Excel Schema Processor", font=("Arial", 16, "bold"))
title_label.pack(pady=10)

Schema_file_path = ""
output_dir = ""

Schema_button = Button(root, text="Select Schema File", command=select_ref_file)
Schema_button.pack(pady=5)
Schema_label = Label(root, text="Schema: Not selected", wraplength=400)
Schema_label.pack()


output_button = Button(root, text="Select Output Folder", command=select_output_dir)
output_button.pack(pady=5)
output_label = Label(root, text="Output folder: Not selected", wraplength=400)
output_label.pack()

# compare_button = Button(root, text="Comparer", command=main_separator)
# compare_button.pack(pady=20)

compare_button = Button(root, text="Process Files", command=main_separator, 
                      bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), 
                      width=15, height=1)
compare_button.pack(pady=20)


status_label = Label(root, text="", fg="black")
status_label.pack()

tk.mainloop()