import tkinter as tk
from tkinter import filedialog, Label, Button, messagebox
from PIL import Image, ImageTk
import re
import os
import sys
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import time
import threading
import queue
from tkinter import simpledialog

pd.set_option('future.no_silent_downcasting', True)


def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

color_dict = {
    'Y': 'FFFF00',  # Yellow
    'G': '00FF00',  # Green
    'L': '0000FF',  # Blue
    'R': 'FF0000',  # Red
    'W': 'FFFFFF',  # White
    'LG': '808080', # Gray (Light Gray)
    'O': 'FFA500',  # Orange
    'BR': '80471c',#'A52A2A', # Brown
    'V': '800080',#'8A2BE2',   # Violet
    'GY': '707070',#'C0C0C0',  # Gray
    'B': '000000' ,  # Black (Noir)
    'P' : 'FFC0CB', # Pink
    'C' : '00FFFF', # Cyan
    'D' : 'FFFFF0', # Ivory
    'SI' : '505050', # Silver
}

def apply_color(symbol, cell):
    if '/' in str(symbol):
        primary, secondary = symbol.split('/')
        fill = PatternFill(start_color=color_dict[primary], end_color=color_dict[primary], fill_type="solid")
        cell.fill = fill
        
        side = Side(border_style="thick", color=color_dict[secondary])
        border = Border(diagonal=side, diagonalUp=True)
        cell.border = border
    elif symbol in color_dict:
        fill = PatternFill(start_color=color_dict[symbol], end_color=color_dict[symbol], fill_type="solid")
        cell.fill = fill
    else :
        cell.value = symbol
    

def header_final_schema(file_path, status_queue=None):
    try:
        wb = load_workbook(file_path)
        for ws in wb.worksheets:
            column_widths = [15, 20, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            thin_border = Border(
                left=Side(style='medium'), 
                right=Side(style='medium'), 
                top=Side(style='medium'), 
                bottom=Side(style='medium')
            )
            
            # Header row 1 - Yazaki logo area
            logo_path = resource_path("yazaki_logo.png")
            if os.path.exists(logo_path):
                try:
                    img = Image(logo_path)
                    ws.merge_cells('A1:J1')
                    ws.add_image(img, 'A2')
                except Exception as e:
                    print(f"Could not add logo: {e}")
                    ws['A1'] = 'YAZAKI'
                    ws['A1'].font = Font(name='Arial', size=14, bold=True)
                    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
            else:
                print(f"Logo file not found at {logo_path}")
                ws['A1'] = 'YAZAKI'
                ws['A1'].font = Font(name='Arial', size=14, bold=True)
                ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
            
            blue_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
            blue_font = Font(name='Arial', size=10, color=Color(rgb='FFFFFF'))
            
            ws.merge_cells('K1:O1')
            first_cell = ws.cell(row=1, column=11)  
            first_cell.value = 'PU24/PU25 LHD'
            first_cell.font = blue_font
            first_cell.fill = blue_fill
            first_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            header_font = Font(name='Arial', size=10)
            
            cell_configs = [
                ('C2', 'Nº du Produit / Niveau', header_font, header_fill, thin_border),
                ('C3', 'KAR G60', Font(name='Arial', size=10), None, thin_border),
                ('E2:F2', 'Local du travail', header_font, header_fill, thin_border),
                ('E3:F3', f'{ws.title}', Font(name='Arial', size=10), None, thin_border),
                ('G2:H2', 'N° de ligne', header_font, header_fill, thin_border),
                ('G3:H3', '1', Font(name='Arial', size=10), None, thin_border),
                ('I2:J2', 'Processus', header_font, header_fill, thin_border),
                ('I3:J3', 'C2', Font(name='Arial', size=10), None, thin_border),
                ('K2:N2', 'Nº de Registre', header_font, header_fill, thin_border),
                ('K3:N3', 'EA-EN-MMO-xx-T-6047', Font(name='Arial', size=10), None, thin_border)
            ]
            
            for cell_range, value, font, fill, border in cell_configs:
                if ':' in cell_range:
                    ws.merge_cells(cell_range)
                    # Get the first cell of the merged range
                    first_cell = ws[cell_range.split(':')[0]]
                    first_cell.value = value
                    first_cell.font = font
                    
                    if fill:
                        first_cell.fill = fill
                    
                    
                    for merged_range in ws.merged_cells.ranges:
                        for row in ws[merged_range.coord]:
                            for cell in row:
                                cell.border = border
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell = ws[cell_range]
                    cell.value = value
                    cell.font = font
                    
                    if fill:
                        cell.fill = fill
                    
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add borders to entire merged ranges
            border_ranges = ['C2:D2', 'I2:J2', 'M2:O2', 'C3:D3', 'I3:J3', 'M3:O3']
            for cell_range in border_ranges:
                for row in ws[cell_range]:
                    for cell in row:
                        cell.border = thin_border
        
        wb.save(file_path)
        print("Header created successfully!")
        if status_queue:
            status_queue.put(("header", True))
        return True
    except Exception as e:
        print(f"Error in header_final_schema: {e}")
        if status_queue:
            status_queue.put(("header", False, str(e)))
        return False

def schema_final_without_color(path_output_excel, df_input, all_sheets, status_queue=None):
    try:
        with pd.ExcelWriter(path_output_excel, engine='xlsxwriter') as writer:      
            for sheet in all_sheets:
                if pd.isna(sheet):
                    continue
                
                sps_data = df_input[df_input['sps'] == sheet]
                
                if sps_data.empty:
                    continue
                
                pm = sps_data['production module ma15'].fillna('').astype(str)
                step = sps_data['step'].fillna('').infer_objects(copy=False).astype(str)
                
                arrays = [np.array(pm), np.array(step)]
                index = pd.MultiIndex.from_arrays(arrays, names=('production module ma15', 'sequence'))
                
                # Création du DataFrame avec la colonne renommée
                df = pd.DataFrame({'Materiel': sps_data['component name'].values,
                                'SAP NO MA15': sps_data['sap no ma15'].values,
                                'Note': sps_data['note'].values,
                                'CS' : sps_data['cs'].values,
                                'color' : sps_data['colour'].values,
                                '' : '',
                                'CON A' : sps_data['from connector'].values,
                                'CAV A' : sps_data['from cavity'].values,
                                'INSERTION A' : '',
                                'CON B' : sps_data['to connector'].values,
                                'CAV B' : sps_data['to cavity'].values,
                                'INSERTION B' : '',
                                },
                                index=index)
                
                sheet = sheet.replace('/', ' ')

                df.to_excel(writer, sheet_name=sheet, startrow=4, startcol=0, index=True)
        print('Schema without color created successfully!')
        if status_queue:
            status_queue.put(("schema_no_color", True))
        return True
    except Exception as e:
        print(f"Error in schema_final_without_color: {e}")
        if status_queue:
            status_queue.put(("schema_no_color", False, str(e)))
        return False


def add_unique_ids_to_schema(wb, sheet_name):
    ws = wb[sheet_name]
        
    header_row = 5
        
    ws.insert_cols(1)
    id_cell = ws.cell(row=header_row, column=1)
    id_cell.value = "Séquence PM"
    id_cell.font = Font(bold=True)
    id_cell.alignment = Alignment(horizontal='center', vertical='center')
        
    id_counter = 1
    for row in range(header_row + 1, ws.max_row + 1):
        has_content = any(ws.cell(row=row, column=col).value for col in range(2, ws.max_column + 1))
        if has_content:
            id_cell = ws.cell(row=row, column=1)
            id_cell.value = id_counter
            id_cell.alignment = Alignment(horizontal='center', vertical='center')
            id_counter += 1
    
    return ws

def schema_final_with_color(wb_path, status_queue=None):
    try:
        wb = load_workbook(wb_path)

        for sheet_name in wb.sheetnames:
            sheet = add_unique_ids_to_schema(wb, sheet_name)
            
            color_col = None
            header_row = 5
            
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=header_row, column=col).value == "color":
                    color_col = col
                    break
            
            if color_col is None:
                print(f"No color column found in sheet: {sheet_name}")
                continue
            
            for row in range(header_row + 1, sheet.max_row + 1):
                color_symbol = sheet.cell(row=row, column=color_col).value
                if color_symbol:
                    cell = sheet.cell(row=row, column=color_col + 1)
                    try:
                        apply_color(color_symbol, cell)
                    except KeyError:
                        print(f"Unknown color code: {color_symbol} in sheet {sheet_name}, row {row}")
                        cell.value = "Unknown color code"
        
        print("Color formatting completed successfully.")
        wb.save(wb_path)
        if status_queue:
            status_queue.put(("schema_color", True))
        return True
    except Exception as e:
        print(f"Error in schema_final_with_color: {e}")
        if status_queue:
            status_queue.put(("schema_color", False, str(e)))
        return False

def add_last_table(file_path, status_queue=None):
    try:
        print('add last table')
        wb = load_workbook(file_path)
        
        for ws in wb.worksheets:
            row_start = ws.max_row + 2  
            table_data = [
                ["", "PM Basique", "", "Niveau", "N° de Phase", "Date", "Préparé par", "Timbre"],
                ["Note:", "Les cases colorées sont des PM optionnelles", "", "", "", "", "", ""],
                ["", "🌀 : ", "A Insérer", "", "", "", "", ""],
                ["", "Ø : ", "A Ne pas insérer", "", "", "", "", ""],
                ["", "⊖ : ", "Déjà inséré", "", "", "", "", ""]
            ]

            # Définir les bordures
            border = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )

            for row_index, row_data in enumerate(table_data, start=row_start):
                for col_index, cell_value in enumerate(row_data, start=1):
                    ws.cell(row=row_index, column=col_index, value=cell_value)
                    ws.cell(row=row_index, column=col_index).border = border
                    ws.cell(row=row_index, column=col_index).font = Font(size=12, bold=True)
                
        print('done...')        
        wb.save(file_path)
        if status_queue:
            status_queue.put(("last_table", True))
        return True
    except Exception as e:
        print(f"Error in add_last_table: {e}")
        if status_queue:
            status_queue.put(("last_table", False, str(e)))
        return False

class ExcelProcessor:
    def __init__(self, root):
        self.root = root
        self.Schema_file_path = ""
        self.output_dir = ""
        self.processing = False
        self.status_queue = queue.Queue()
        
        self.setup_ui()
        self.check_queue()
    
    def setup_ui(self):
        title_label = Label(self.root, text="Excel Schema Processor", font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        designer_label = tk.Label(
                                root,
                                text="Designed by Abdelhamid Chebel & Ismail Zahraoui",
                                font=("Helvetica", 8),
                                fg="gray"
                            )
        designer_label.pack(side="bottom", pady=5)
        
        self.Schema_button = Button(self.root, text="Select Schema File", command=self.select_ref_file)
        self.Schema_button.pack(pady=5)
        self.Schema_label = Label(self.root, text="Schema: Not selected", wraplength=400)
        self.Schema_label.pack()
        
        self.output_button = Button(self.root, text="Select Output Folder", command=self.select_output_dir)
        self.output_button.pack(pady=5)
        self.output_label = Label(self.root, text="Output folder: Not selected", wraplength=400)
        self.output_label.pack()
        
        self.process_button = Button(self.root, text="Process Files", command=self.main_separator, 
                              bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), 
                              width=15, height=1)
        self.process_button.pack(pady=20)
        
        # Add a button to show contact information again
        self.contact_button = Button(self.root, text="Contact Info", command=self.show_contact_info, 
                                     bg="#2196F3", fg="white", font=("Arial", 10, "bold"), 
                                     width=15, height=1)
        self.contact_button.pack(pady=10)
        
        # Add a button to show rules for using the app
        self.rules_button = Button(self.root, text="App Rules", command=self.show_app_rules, 
                                   bg="#FF9800", fg="white", font=("Arial", 10, "bold"), 
                                   width=15, height=1)
        self.rules_button.pack(pady=10)
        
        self.progress_label = Label(self.root, text="", fg="blue")
        self.progress_label.pack()
        
        self.status_label = Label(self.root, text="", fg="black")
        self.status_label.pack()

    def show_contact_info(self):
        email = "abdelhamid.chebel04@gmail.com"
        self.root.clipboard_clear()
        self.root.clipboard_append(email)
        self.root.update()  # Keep the clipboard content
        messagebox.showinfo(
            "Contact Information",
            f"If you have any problems, contact me at {email}\n"
            "and place in the subject: 'Feedback and Issues – YMM-2 App Schema Tool'.\n\n"
            "The email address has been copied to your clipboard."
        )

    def show_app_rules(self):
        messagebox.showinfo(
            "Important Information",
            "Please ensure the sheet is named 'Principal' and contains the following columns:\n\n"
            "- SPS\n"
            "- Production Module MA15\n"
            "- Step\n"
            "- Component Name\n"
            "- SAP NO MA15\n"
            "- Note\n"
            "- CS\n"
            "- Colour\n"
            "- FROM CONNECTOR\n"
            "- FROM CAVITY\n"
            "- TO CONNECTOR\n"
            "- TO CAVITY"
        )
    
    def select_ref_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls;*.XLSX")])
        
        if not file_path:
            return
            
        if not file_path.lower().endswith((".xlsx", ".xls")):
            messagebox.showerror("Erreur", "Please select a valid Excel file (*.xlsx, *.xls;*.XLSX)")
            return

        self.Schema_file_path = file_path
        self.Schema_label.config(text=f"Schema File: {os.path.basename(self.Schema_file_path)}")
    
    def select_output_dir(self):
        output_dir = filedialog.askdirectory()
        if output_dir:
            self.output_dir = output_dir
            self.output_label.config(text=f"Output folder: {os.path.basename(self.output_dir)}")
    
    def check_queue(self):
        try:
            while not self.status_queue.empty():
                message = self.status_queue.get(0)
                
                if message[0] == "progress":
                    self.progress_label.config(text=f"Processing: {message[1]}")
                elif message[0] == "complete":
                    success = message[1]
                    output_file = message[2]
                    self.processing = False
                    self.process_button.config(state="normal")
                    
                    if success:
                        self.status_label.config(text=f"Success! File saved as {os.path.basename(output_file)}", fg="green")
                        self.progress_label.config(text="")
                        messagebox.showinfo("Success", "File processing completed successfully!")
                    else:
                        error = message[3] if len(message) > 3 else "Unknown error"
                        self.status_label.config(text=f"Processing failed: {error}", fg="red")
                        self.progress_label.config(text="")
                elif message[0] in ["header", "schema_no_color", "schema_color", "last_table"]:
                    success = message[1]
                    step_name = {
                        "header": "Creating header",
                        "schema_no_color": "Creating base schema",
                        "schema_color": "Adding colors",
                        "last_table": "Adding final table"
                    }
                    if success:
                        self.progress_label.config(text=f"Completed: {step_name[message[0]]}")
                    else:
                        error = message[2] if len(message) > 2 else "Unknown error"
                        self.progress_label.config(text=f"Failed: {step_name[message[0]]} - {error}")
        except Exception as e:
            print(f"Error in check_queue: {e}")
        
        # Check again after 100ms
        self.root.after(100, self.check_queue)
    
    def process_thread(self, Schema_file_path, output_file):
        try:
            self.status_queue.put(("progress", "Loading data..."))
            df_input = pd.read_excel(Schema_file_path, sheet_name="Principal")
            print(df_input.columns)
            print('----------------')
            # df_input =  df_input.select_dtypes(exclude=['float64'])
            df_input.columns = df_input.columns.str.lower()
            print(df_input.columns)
            # print(df_input['sps'].unique())
            df_input['sps'] = df_input['sps'].apply(lambda x: x.lower() if isinstance(x, str) else x)
            all_sheets = list(df_input['sps'].unique())
            print(all_sheets)

            if not all_sheets:
                self.status_queue.put(("complete", False, output_file, "No valid SPS values found in the input file"))
                return
            
            self.status_queue.put(("progress", "Creating base schema..."))
            if not schema_final_without_color(output_file, df_input, all_sheets, self.status_queue):
                self.status_queue.put(("complete", False, output_file, "Failed to create base schema"))
                return
            
            self.status_queue.put(("progress", "Adding header..."))
            if not header_final_schema(output_file, self.status_queue):
                self.status_queue.put(("complete", False, output_file, "Failed to create header"))
                return
            
            self.status_queue.put(("progress", "Adding colors..."))
            if not schema_final_with_color(output_file, self.status_queue):
                self.status_queue.put(("complete", False, output_file, "Failed to add colors"))
                return
            
            self.status_queue.put(("progress", "Adding final table..."))
            if not add_last_table(output_file, self.status_queue):
                self.status_queue.put(("complete", False, output_file, "Failed to add final table"))
                return
            
            self.status_queue.put(("progress", "Finalizing document..."))
            if not header_final_schema(output_file, self.status_queue):
                self.status_queue.put(("complete", False, output_file, "Failed to finalize document"))
                return
            
            print(f"Workbook saved successfully to {output_file}")
            self.status_queue.put(("complete", True, output_file))
            os.startfile(output_file)
        except Exception as e:
            print(f"Error in process_thread: {e}")
            self.status_queue.put(("complete", False, output_file, str(e)))
    
    def main_separator(self):
        if not self.Schema_file_path or not self.output_dir:
            messagebox.showerror("Erreur", "Please select both Schema file and output folder")
            return
        
        # Add pop-up to inform the user about the required sheet name and columns
        messagebox.showinfo(
            "Important Information",
            "Please ensure the sheet is named 'Principal' and contains the following columns:\n"
            "SPS, Production Module MA15, Step, Component Name, SAP NO MA15, Note, CS, Colour, "
            "FROM CONNECTOR, FROM CAVITY, TO CONNECTOR, TO CAVITY"
        )
        
        if self.processing:
            messagebox.showinfo("Info", "Processing already in progress. Please wait.")
            return
        
        try:
            timestamp = time.strftime("%Y%m%d-%H%M%S")
            output_file = os.path.join(self.output_dir, f"Final_Schema_{timestamp}.xlsx")
            
            self.processing = True
            self.process_button.config(state=tk.DISABLED)
            self.status_label.config(text="Processing...", fg="blue")
            
            # Start the processing in a separate thread
            processing_thread = threading.Thread(
                target=self.process_thread,
                args=(self.Schema_file_path, output_file)
            )
            processing_thread.daemon = True
            processing_thread.start()
            
            
        except Exception as e:
            self.processing = False
            self.process_button.config(state="normal")
            messagebox.showerror("Erreur", f"Error starting processing: {e}")


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Excel Schema Processor - YMM-2")
    
    window_width = 400
    window_height = 450
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x_position = (screen_width - window_width) // 2
    y_position = (screen_height - window_height) // 2
    root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
    
    app = ExcelProcessor(root)

    root.mainloop()

