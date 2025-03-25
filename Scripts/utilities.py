import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

color_dict = {
    'Y': 'FFFF00',  # Yellow
    'G': '00FF00',  # Green
    'L': '0000FF',  # Blue
    'R': 'FF0000',  # Red
    'W': 'FFFFFF',  # White
    'LG': '808080', # Gray (Light Gray)
    'O': 'FFA500',  # Orange
    'BR': '80471c',#'A52A2A', # Brown
    'V': '800080',#'8A2BE2',   # Violet
    'GY': '707070',#'C0C0C0',  # Gray
    'B': '000000'   # Black (Noir)
}

def apply_color(symbol, cell):
    if '/' in symbol:
        primary, secondary = symbol.split('/')
        fill = PatternFill(start_color=color_dict[primary], end_color=color_dict[primary], fill_type="solid")
        cell.fill = fill
        
        side = Side(border_style="thick", color=color_dict[secondary])
        border = Border(diagonal=side, diagonalUp=True)
        cell.border = border
    else:
        fill = PatternFill(start_color=color_dict[symbol], end_color=color_dict[symbol], fill_type="solid")
        cell.fill = fill

# Créer un nouveau classeur et une feuille de calcul
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Colors"

# En-têtes des colonnes
sheet.append(["Symboles", "Couleur"])

data = ["BR", "Y", "R/W", "W/R", "G/L", "O/W"]

# Remplissage des données et application des couleurs
row_num = 2  # Première ligne après les en-têtes
for symbol in data:
    sheet.cell(row=row_num, column=1, value=symbol)
    cell = sheet.cell(row=row_num, column=2, value=symbol)
    apply_color(symbol, cell)
    row_num += 1

# Ajuster la largeur des colonnes
for col in range(1, 3):
    sheet.column_dimensions[get_column_letter(col)].width = 15

wb.save("colored_excel.xlsx")
print("Fichier Excel généré avec succès !")